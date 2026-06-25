"""
从 xlsx 导入 2023-2025 年广东物理类本科批录取数据到 colleges.csv（带 college_province 列）。
支持原始项目下含 rank_lookup.py 的结构。
"""

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

XLSX_PATH = Path("/Users/csr/Documents/2023-25年近三年广东物理类本科批录取分数线.xlsx")
OUTPUT_PATH = Path(__file__).parent.parent / "app" / "data" / "colleges.csv"


def normalize_program_group(raw: str) -> str:
    raw = raw.strip().strip("（）()")
    if raw.startswith("专业组"):
        raw = raw[3:]
    return raw.strip()


def normalize_school_name(name: str) -> str:
    name = name.strip()
    name = re.sub(r"\s+", "", name)
    return name


def infer_school_type(name: str) -> str:
    if any(kw in name for kw in ("学院", "大学")):
        known_private = (
            "培正", "新华", "华联", "南方", "广州城市理工", "东莞城市",
            "广州华立", "广州华商", "广州理工学院", "珠海科技",
            "广州工商", "广东科技", "广东理工", "广东白云",
            "电子科技大学中山学院", "北京理工大学珠海学院",
            "华南农业大学珠江学院", "广东外语外贸大学南国商学院",
            "湛江科技学院", "广州软件学院"
        )
        if any(kw in name for kw in known_private):
            return "民办本科"
        return "公办本科"
    return "其他"


def parse_sheet_25(ws) -> dict:
    """解析25年Sheet: 广东省考试院投档汇总"""
    data = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        if not row[0] or not isinstance(row[0], (int, float)):
            continue
        code = str(int(row[0]))
        name = normalize_school_name(str(row[1] or ""))
        group = normalize_program_group(str(row[2] or ""))
        score = int(row[5]) if row[5] else 0
        rank = int(row[6]) if row[6] else 0
        key = (code, group)
        data[key] = {
            "college_code": code,
            "college_name": name,
            "program_group": group,
            "min_score": score,
            "min_rank": rank,
            "year": 2025,
        }
    return data


def parse_sheet_23_or_24(ws, year: int, has_city: bool = False) -> dict:
    """解析23年/24年Sheet（专业级 → 聚合到专业组级）"""
    grouped = defaultdict(list)
    metadata = {}

    for row in list(ws.iter_rows(values_only=True)):
        if not row[0] or row[0] == "生源地":
            continue
        school = normalize_school_name(str(row[1] or ""))
        group = normalize_program_group(str(row[4] or ""))
        subject = str(row[6] or "").strip()
        score = row[7]
        rank_val = row[8]

        if score is None or rank_val is None:
            continue
        try:
            score = int(score)
            rank_val = int(rank_val)
        except (ValueError, TypeError):
            continue

        key = (school, group)
        grouped[key].append((score, rank_val))

        if key not in metadata:
            info = {"subject_requirement": subject}
            if has_city and len(row) > 11:
                info["city"] = str(row[11] or "").strip()
                info["province"] = str(row[10] or "").strip()
            metadata[key] = info

    result = {}
    for (school, group), scores in grouped.items():
        min_score = min(s[0] for s in scores)
        min_rank = min(s[1] for s in scores)
        meta = metadata.get((school, group), {})
        result[(school, group)] = {
            "college_name": school,
            "program_group": group,
            "min_score": min_score,
            "min_rank": min_rank,
            "year": year,
            "subject_requirement": meta.get("subject_requirement", ""),
            "city": meta.get("city", ""),
            "province": meta.get("province", ""),
        }
    return result


def merge_data(data_23: dict, data_24: dict, data_25: dict) -> list[dict]:
    """合并三个年份数据，为每个记录添加 college_province（院校所在省份）"""
    name_to_codes = defaultdict(set)
    for key, rec in data_25.items():
        name_to_codes[rec["college_name"]].add(rec["college_code"])

    name_to_pg_codes = defaultdict(list)
    for (code, pg), rec in data_25.items():
        name_to_pg_codes[rec["college_name"]].append((code, pg))

    # 从24年Sheet提取city/province元数据（院校所在地）
    school_meta = defaultdict(dict)
    for key, rec in data_24.items():
        school = rec["college_name"]
        if rec.get("city"):
            school_meta[school]["city"] = rec["city"]
        if rec.get("province"):
            school_meta[school]["college_province"] = rec["province"]

    pg_subject_req = {}
    for key, rec in {**{(k, v["program_group"]): v for k, v in data_23.items()},
                     **{(k, v["program_group"]): v for k, v in data_24.items()}}.items():
        pg_subject_req[key] = rec.get("subject_requirement", "")

    output_rows = []
    seen = set()

    # 25年（含院校代码）
    for (code, pg), rec_25 in data_25.items():
        school = rec_25["college_name"]
        meta = school_meta.get(school, {})

        subject_req = pg_subject_req.get((school, pg), "")
        if not subject_req:
            for key, rec24 in data_24.items():
                if rec24["college_name"] == school:
                    subject_req = rec24.get("subject_requirement", "物理+不限")
                    break
        if not subject_req:
            subject_req = "物理+不限"

        school_type = infer_school_type(school)
        city = meta.get("city", "")
        college_province = meta.get("college_province", "")

        row_key = (code, pg, 2025)
        if row_key in seen:
            continue
        seen.add(row_key)

        output_rows.append({
            "province": "广东",
            "subject_group": "物理",
            "college_code": code,
            "college_name": school,
            "school_type": school_type,
            "college_province": college_province,
            "city": city,
            "program_group": pg,
            "subject_requirement": subject_req,
            "year": 2025,
            "min_score": rec_25["min_score"],
            "min_rank": rec_25["min_rank"],
        })

    # 23年 和 24年
    for year_data, year in [(data_23, 2023), (data_24, 2024)]:
        for key, rec in year_data.items():
            school = rec["college_name"]
            codes = name_to_codes.get(school, set())
            if not codes:
                for name in name_to_codes:
                    if school in name or name in school:
                        codes = name_to_codes[name]
                        break
                if not codes:
                    codes = {"00000"}

            code = list(codes)[0]
            pg = rec["program_group"]
            city = rec.get("city", "") or school_meta.get(school, {}).get("city", "")
            college_province = rec.get("province", "") or school_meta.get(school, {}).get("college_province", "")

            subject_req = rec.get("subject_requirement", "")
            if not subject_req:
                subject_req = pg_subject_req.get((school, pg), "物理+不限")
            if not subject_req:
                subject_req = "物理+不限"

            school_type = infer_school_type(school)
            row_key = (code, pg, year)
            if row_key in seen:
                continue
            seen.add(row_key)

            output_rows.append({
                "province": "广东",
                "subject_group": "物理",
                "college_code": code,
                "college_name": school,
                "school_type": school_type,
                "college_province": college_province,
                "city": city,
                "program_group": pg,
                "subject_requirement": subject_req,
                "year": year,
                "min_score": rec["min_score"],
                "min_rank": rec["min_rank"],
            })

    return output_rows


def main():
    print(f"读取 xlsx: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)

    print("解析 23年 Sheet...")
    data_23 = parse_sheet_23_or_24(wb["23年"], 2023, has_city=False)
    print(f"  23年: {len(data_23)} 条")

    print("解析 24年 Sheet...")
    data_24 = parse_sheet_23_or_24(wb["24年"], 2024, has_city=True)
    print(f"  24年: {len(data_24)} 条")

    print("解析 25年 Sheet...")
    data_25 = parse_sheet_25(wb["25年"])
    print(f"  25年: {len(data_25)} 条")

    print("合并数据...")
    rows = merge_data(data_23, data_24, data_25)
    print(f"  合计: {len(rows)} 条")

    # 去重
    seen = set()
    unique_rows = []
    for row in rows:
        key = (row["college_code"], row["program_group"], row["year"])
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)
    print(f"  去重后: {len(unique_rows)} 条")

    unique_rows.sort(key=lambda r: (r["college_code"], r["program_group"], r["year"]))

    fieldnames = [
        "province", "subject_group", "college_code", "college_name",
        "school_type", "college_province", "city", "program_group",
        "subject_requirement", "year", "min_score", "min_rank",
    ]

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(unique_rows)

    print(f"\n写入完成: {OUTPUT_PATH}")
    print(f"  总行数: {len(unique_rows)}")

    schools = set(r["college_code"] for r in unique_rows)
    years = set(r["year"] for r in unique_rows)
    print(f"  院校数: {len(schools)}")
    print(f"  年份: {sorted(years)}")

    from collections import Counter
    year_counts = Counter(r["year"] for r in unique_rows)
    for y in sorted(year_counts):
        print(f"  {y}年: {year_counts[y]} 条")


if __name__ == "__main__":
    main()
